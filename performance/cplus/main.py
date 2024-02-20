"""
drawing api perf test
"""

import os
import re
import subprocess
import threading
import time
import logging
from ctypes import cdll
import json
from statistics import mean
import pandas
from openpyxl.utils import get_column_letter
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Side, Border, Font, colors
# import colorlog
import datetime


class Logging(object):
    def log(self, level="INFO"):  # 生成日志的主方法,传入对那些级别及以上的日志进行处理
        log_colors_config = {
            "DEBUG": "white",
            "INFO": "thin",
            "WARNING": "yellow",
            "ERROR": "red",
            "CRITICAL": "bold_red",
        }
        logger = logging.getLogger()  # 创建日志器
        console_handler = logging.StreamHandler()  # 创建控制台日志处理器
        file_name = datetime.datetime.now().strftime('%Y%m%d_%H%M%S') + ".log"
        log_path = os.path.join(os.getcwd(), r"log")
        log_file_path = os.path.join(log_path, file_name)
        os.makedirs(log_path, exist_ok=True)
        file_handler = logging.FileHandler(filename=log_file_path, mode='a', encoding="utf-8")  # 创建日志文件处理器

        logger.setLevel(level)  # logger控制最低输出什么级别日志(优先级最高)
        # console_handler.setLevel(logging.DEBUG) # #console_handler设置控制台最低输出级别日志
        # file_handler.setLevel(logging.INFO) # file_handler设置保存到文件最低输出级别日志

        # 创建格式器
        file_formatter = logging.Formatter(
            fmt='[%(asctime)s.%(msecs)03d] [%(filename)s, line:%(lineno)d] [%(levelname)s]: %(message)s',
            datefmt='%Y-%m-%d  %H:%M:%S')

        # console_formatter = colorlog.ColoredFormatter(
        #     fmt='%(log_color)s[%(asctime)s.%(msecs)03d] [%(filename)s, line:%(lineno)d] [%(levelname)s]: %(message)s',
        #     datefmt='%Y-%m-%d  %H:%M:%S',
        #     log_colors=log_colors_config)
        # 给处理器添加格式
        # console_handler.setFormatter(fmt=console_formatter)
        file_handler.setFormatter(fmt=file_formatter)

        if not logger.handlers:  # 这里进行判断, 如果logger.handlers列表为空, 则添加, 否则, 直接去写日志, 解决重复打印的问题
            logger.addHandler(console_handler)
            logger.addHandler(file_handler)
        return logger  # 返回日志器


logger = Logging().log(level="INFO")


class ExecuteHDC:
    def __init__(self, hdc, device_id):
        self.hdc = hdc
        self.device_id = device_id

    # 执行命令，不带有返回结果
    def execute_command(self, command, shell=True):
        command = self.get_command(command, shell)
        logger.info("command: %s", command)
        # 执行命令
        poc = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, shell=shell)
        out, error = poc.communicate(timeout=60)
        logger.info(out)
        logger.info("The command is executed successfully")
        if error:
            logger.error(error)

    # 执行命令，带有返回结果
    def execute_command_result(self, command, shell=True):
        command = self.get_command(command, shell)
        logger.info("command: %s", command)
        result = subprocess.getoutput(command)
        logger.info(result)
        return result

    # 获取命令
    def get_command(self, command, shell=True):
        if shell:
            if self.device_id:
                command = '"' + self.hdc + '"' + " -t " + self.device_id + " shell " + command
            else:
                command = '"' + self.hdc + '"' + " shell " + command
        else:
            if self.device_id:
                command = '"' + self.hdc + '"' + " -t " + self.device_id + " " + command
            else:
                command = '"' + self.hdc + '"' + " " + command

        return command


class MyThread(threading.Thread):
    def __init__(self, func, args=()):
        super(MyThread, self).__init__()
        self.result = None
        self.func = func
        self.args = args

    def run(self):
        self.result = self.func(*self.args)

    def get_result(self):
        try:
            # 如果子线程不使用join方法, 此处可能会报没有self.result的错误
            return self.result
        except Exception:
            return None


def get_my_path():
    my_path = os.path.dirname(os.path.realpath(__file__))
    return my_path


# 启动rs demo
def push_and_run_rs_demo(hdcServer, file_path, api_name, test_count, test_time):
    file_name = file_path.split(os.sep)[-1]
    command_list = [
        "mount -o rw,remount /",
        "power-shell setmode 602",
        "power-shell display -s 50",
        f"file send {file_path} /data/local/tmp",
        f"chmod +x /data/local/tmp/{file_name}",
        f"/data/local/tmp/{file_name} {api_name} {test_count} {test_time}",
    ]
    for command in command_list:
        if "file" in command:
            ret = hdcServer.execute_command_result(command, shell=False)
            logger.info(f"ret: {ret}")
        else:
            ret = hdcServer.execute_command_result(command, shell=True)
            logger.info(f"ret: {ret}")


# 获取应用进程号
def get_pid(hdcServer, name):
    command = "ps -ef | findstr {0}".format(name)
    result = hdcServer.execute_command_result(command, shell=True)
    logger.info(f"result is :{result}")
    pid = re.findall(r"([\d.]+)", result)[0]
    logger.info(f"{name} pid is : %s", pid)

    return pid


# 杀死进程
def kill_pid(hdcServer, pid):
    hdcServer.execute_command("kill -39 {0}".format(pid), shell=True)
    return True


# 开启无障碍开关(arkui组件调试开关), 获取一些UI控件信息
def start_uitest(hdcServer, wait_time=60):
    command_list = [
        "mount -o remount,rw /",
        "param set persist.ace.testmode.enabled 1",
        "param set persist.ace.debug.enabled 1",
        "reboot"
    ]
    for command in command_list:
        hdcServer.execute_command(command, shell=True)
    time.sleep(wait_time)

    return True


def unlock(hdcServer):
    # yellow todo
    # start_x, start_y, end_x, end_y, swipe_speed = 1320, 2400, 1320, 2000, 200
    start_x, start_y, end_x, end_y, swipe_speed = 350, 1000, 350, 600, 200
    command_list = [
        f"power-shell wakeup",
        f"uinput -T -m {start_x} {start_y} {end_x} {end_y} {swipe_speed}",
        f"power-shell setmode 602",
        f"power-shell display -s 50",
    ]
    for command in command_list:
        hdcServer.execute_command(command, shell=True)
    time.sleep(2)


def reboot_and_unlock(hdcServer, wait_time=60):
    start_uitest(hdcServer, wait_time=wait_time)
    unlock(hdcServer)


def set_hilog(hdcServer):
    command_list = [
        "hilog -r",
        "hilog -b D",
        "hilog -G 512M",
        "hilog -Q pidoff",
        "hilog -p off",
    ]
    for command in command_list:
        hdcServer.execute_command_result(command, shell=True)


def delete_hilog(hdcServer):
    delete_cmd = "rm -rf /data/log/hilog/*"
    hdcServer.execute_command_result(delete_cmd, shell=True)


def get_hilog_file_name(hdcServer, api_name):
    file_name = ""
    query_cmd = "ls /data/log/hilog"
    result = hdcServer.execute_command_result(query_cmd, shell=True).split("\n")
    for file in result:
        if api_name in file:
            file_name = file
            break
    logger.info(f"{api_name} hilog file name is: {file_name.strip()}")
    return file_name.strip()


def make_dir(dir_path):
    os.makedirs(dir_path, exist_ok=True)
    return dir_path


def file_valid(filename):
    if os.path.exists(filename) and os.path.getsize(filename) > 0:
        return True
    return False


def collecte_hilog(hdcServer, api_name, log_name, mode="drawing", wait_time=10):
    file_dir = os.path.dirname(os.path.realpath(__file__))
    save_path = os.path.join(file_dir, "reports", mode, api_name)
    make_dir(save_path)
    hdcServer.execute_command_result("hilog -r", shell=True)
    hdcServer.execute_command_result("mount -o rw,remount /", shell=True)
    hdcServer.execute_command_result("hilog -w stop", shell=True)
    delete_hilog(hdcServer)
    time.sleep(1)
    # 日志落盘
    hilog_start_cmd = f"hilog -w start -f {api_name} -m none"
    hdcServer.execute_command_result(hilog_start_cmd, shell=True)
    time.sleep(wait_time)

    # 停止日志落盘
    hilog_stop_cmd = "hilog -w stop"
    hdcServer.execute_command_result(hilog_stop_cmd, shell=True)
    file_name = get_hilog_file_name(hdcServer, api_name)
    logger.info(f"file name is: {file_name}")
    pull_cmd = f"file recv /data/log/hilog/{file_name} {save_path}/{log_name}.log"
    hdcServer.execute_command_result(pull_cmd, shell=False)
    hilog_file_path = os.path.join(save_path, f"{log_name}.log")
    # 删除hilog日志
    delete_hilog(hdcServer)
    return hilog_file_path


def clear_soc_perf(hdcServer, clear_flag=False, wait_time=30):
    # yellow todo
    # command_list = [
    #     "mount -o rw,remount /",
    #     "mount /sys_prod -o remount,rw",
    #     "file recv /sys_prod/etc/soc_perf ./",
    #     "truncate -s 0 /sys_prod/etc/soc_perf/*.xml",
    #     "reboot",
    # ]
    command_list = [
        "mount -o rw,remount /",
        "mount /vendor -o remount,rw",
        "file recv /vendor/etc/soc_perf ./",
        "truncate -s 0 /vendor/etc/soc_perf/*.xml",
        "reboot",
    ]
    if clear_flag:
        for command in command_list:
            if command == "reboot":
                hdcServer.execute_command_result(command, shell=True)
                time.sleep(wait_time)
                unlock(hdcServer)
            elif "file" in command:
                hdcServer.execute_command_result(command, shell=False)
            else:
                hdcServer.execute_command_result(command, shell=True)


# 锁频锁核
def lock_core_frequency(hdcServer, clear_flag=False, wait_time=30):
    logger.info("ALN手机 CPU 小、大核锁最低频并且关闭小、大核, 中核锁最高频; GPU、DDR锁最高频")
    clear_soc_perf(hdcServer, clear_flag=clear_flag, wait_time=wait_time)
    command_list = [
        '"echo 1992000>  /sys/devices/system/cpu/cpu0/cpufreq/scaling_max_freq"',
        '"echo 1992000>  /sys/devices/system/cpu/cpu0/cpufreq/scaling_min_freq"',
        '"echo 1992000>  /sys/devices/system/cpu/cpu0/cpufreq/scaling_max_freq"',

        '"echo 1992000>  /sys/devices/system/cpu/cpu1/cpufreq/scaling_max_freq"',
        '"echo 1992000>  /sys/devices/system/cpu/cpu1/cpufreq/scaling_min_freq"',
        '"echo 1992000>  /sys/devices/system/cpu/cpu1/cpufreq/scaling_max_freq"',

        '"echo 1992000>  /sys/devices/system/cpu/cpu2/cpufreq/scaling_max_freq"',
        '"echo 1992000>  /sys/devices/system/cpu/cpu2/cpufreq/scaling_min_freq"',
        '"echo 1992000>  /sys/devices/system/cpu/cpu2/cpufreq/scaling_max_freq"',

        '"echo 1992000>  /sys/devices/system/cpu/cpu3/cpufreq/scaling_max_freq"',
        '"echo 1992000>  /sys/devices/system/cpu/cpu3/cpufreq/scaling_min_freq"',
        '"echo 1992000>  /sys/devices/system/cpu/cpu3/cpufreq/scaling_max_freq"',

        '"echo performance > /sys/devices/system/cpu/cpu0/cpufreq/scaling_governor"',
        '"echo performance > /sys/devices/system/cpu/cpu1/cpufreq/scaling_governor"',
        '"echo performance > /sys/devices/system/cpu/cpu2/cpufreq/scaling_governor"',
        '"echo performance > /sys/devices/system/cpu/cpu3/cpufreq/scaling_governor"',

        # '"echo 0 > sys/devices/system/cpu/cpu0/online"',
        '"echo 0 > sys/devices/system/cpu/cpu1/online"',  # rk上只开1个核太慢太慢了，多开一个核数
        '"echo 0 > sys/devices/system/cpu/cpu2/online"',
        '"echo 0 > sys/devices/system/cpu/cpu3/online"',

        '"echo 800000000>  /sys/class/devfreq/fde60000.gpu/max_freq"',
        '"echo 800000000>  /sys/class/devfreq/fde60000.gpu/min_freq"',
        '"echo 800000000>  /sys/class/devfreq/fde60000.gpu/max_freq"',

        # '"echo 3197000000 > /sys/class/devfreq/ddrfreq/max_freq"',
        # '"echo 3197000000 > /sys/class/devfreq/ddrfreq/min_freq"',
        # '"echo 3197000000 > /sys/class/devfreq/ddrfreq/max_freq"',
        # '"echo 3197000000 > /sys/class/devfreq/ddrfreq_up_threshold/max_freq"',
        # '"echo 3197000000 > /sys/class/devfreq/ddrfreq_up_threshold/min_freq"',
        # '"echo 3197000000 > /sys/class/devfreq/ddrfreq_up_threshold/max_freq"',
    ]
    for command in command_list:
        hdcServer.execute_command_result(command, shell=True)


# 计算接口时间
def cal_api_time(api_name, hilog_file_path):  # 依赖demo打印的hilog时间戳计算执行的时间
    keyword = "DrawingApiTest"
    start_keyword = "DrawingApiTest Started:"
    finished_keyword = "DrawingApiTest Finished:"
    total_api_call_count = "DrawingApiTest TotalApiCallCount:"
    if file_valid(hilog_file_path):
        with open(hilog_file_path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()
            result = re.findall(r"{0}.*".format(keyword), content)
            logging.info(result)
            for line in result:
                if start_keyword in line:
                    start_time = int(re.search(r"\d+", line).group())
                if finished_keyword in line:
                    end_time = int(re.search(r"\d+", line).group())
                if total_api_call_count in line:
                    total_count = int(re.search(r"\d+", line).group())
            if start_time and end_time and total_count:
                api_time = round((end_time - start_time) / total_count, 2)
                logger.info(f"{api_name} start time is:{start_time}, end time is:{end_time}, take time: {api_time}μs")
                return api_time, total_count


# https://codehub-y.huawei.com/Fields/Avatar/Graphic/GraphicsInnovation/graphic_innovation/files?ref=ddgr_dev_1117&filePath=foundation%2Fgraphic%2Fgraphic_2d%2Frosen%2Fmodules%2F2d_graphics%2Ftest%2Frs_demo.cpp&isFile=true#L150


def catch_trace(hdcServer, mode, api_name, trace_name, duration=10):
    trace_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "reports", mode, api_name)
    os.makedirs(trace_path, exist_ok=True)
    trace_cmd = f'"bytrace -t {duration} -b 307200 --overwrite ace app ohos zimage zmedia zcamera zaudio ability ' \
                f'window distributeddatamgr graphic sched freq irq sync mmc rpc workq idle disk pagecache memreclaim ' \
                f'binder multimodalinput ark > /data/local/tmp/{trace_name}.ftrace"'
    hdcServer.execute_command_result(trace_cmd, shell=True)
    sed_cmd = f"sed -i '1,2d' /data/local/tmp/{trace_name}.ftrace"
    hdcServer.execute_command_result(sed_cmd, shell=True)
    file_receive_cmd = f"file recv /data/local/tmp/{trace_name}.ftrace {trace_path}"
    hdcServer.execute_command_result(file_receive_cmd, shell=False)


def hiperf_collect(hdcServer, mode, api_name, data_name, test_time):
    logger.info("xyj in hiperf_collect")
    data_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "reports", mode, api_name)
    os.makedirs(data_path, exist_ok=True)
    pid = get_pid(hdcServer, api_name)
    cmd = f"hiperf record -e hw-instructions,hw-cpu-cycles -f 1000 -s dwarf -p {pid} -d {test_time} -o /data/local/tmp/{data_name}_report.data"
    hdcServer.execute_command_result(cmd, shell=True)
    file_receive_cmd = f"file recv /data/local/tmp/{data_name}_report.data {data_path}"
    hdcServer.execute_command_result(file_receive_cmd, shell=False)
    local_path = os.path.join(data_path, f"{data_name}_report.data")
    return local_path


def hiperf_report(perf_data, report_file):
    logger.info("xyj in hiperf_report")
    my_path = get_my_path()
    lib_path = os.path.join(my_path, "hiperf", "bin", "windows", "x86_64", "libhiperf_report.dll")
    cdll.LoadLibrary(lib_path).ReportJson(perf_data.encode("utf-8"), 'json.txt'.encode("utf-8"))
    time.sleep(2)
    with open('json.txt', 'r') as json_file:
        all_json = json_file.read()
    report_html_path = os.path.join(my_path, "hiperf", "report.html")
    with open(report_html_path, 'r', encoding='utf-8') as html_file:
        html_str = html_file.read()
    with open(report_file, 'w', encoding='utf-8') as report_html_file:
        report_html_file.write(html_str + all_json + '</script>'
                                                     ' </body>'
                                                     ' </html>')
    logger.info("save to %s success" % report_file)
    os.remove('json.txt')


def hiperf_parse(mode, perf_data, api_name, html_name):
    logger.info("xyj in hiperf_parse")
    my_path = get_my_path()
    dir_path = os.path.join(my_path, "reports", mode, api_name)
    html_file_path = os.path.join(dir_path, f"{html_name}.html")
    if file_valid(perf_data):
        hiperf_report(perf_data, html_file_path)
        logger.info("xyj in hiperf_parse")
    cycles_keyword = '"hw-cpu-cycles","eventCount":'
    instructions_keyword = '"hw-instructions","eventCount":'
    if file_valid(html_file_path):
        with open(html_file_path, "r", encoding="utf-8") as f:
            content = f.read()
            cycles_count = re.search(r"{0}\d+".format(cycles_keyword), content).group().split(":")[1]
            instructions_count = re.search(r"{0}\d+".format(instructions_keyword), content).group().split(":")[1]
            logger.info(f"{api_name} cycles is: {cycles_count}, instructions is: {instructions_count}")

        return cycles_count, instructions_count


def push_simpleperf(hdcServer):
    my_path = get_my_path()
    simpleperf_path = os.path.join(my_path, "simpleperf", "bin", "android", "arm64", "simpleperf")
    command_list = [
        "mount -o rw,remount /",
        f"file send {simpleperf_path} /data",
        f"chmod +x /data/simpleperf",
    ]
    for command in command_list:
        if "file" in command:
            ret = hdcServer.execute_command_result(command, shell=False)
            logger.info(f"ret: {ret}")
        else:
            ret = hdcServer.execute_command_result(command, shell=True)
            logger.info(f"ret: {ret}")


def simpleperf_collect(hdcServer, mode, demo_path, api_name, data_name, test_time):
    data_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "reports", mode, api_name)
    os.makedirs(data_path, exist_ok=True)
    demo_name = demo_path.split(os.sep)[-1]
    pid = get_pid(hdcServer, demo_name)
    cmd = f"/data/simpleperf record -e instructions,cpu-cycles -f 1000 -g -p {pid} --duration {test_time} -o /data/local/tmp/{data_name}_report.data"
    hdcServer.execute_command_result(cmd, shell=True)
    file_receive_cmd = f"file recv /data/local/tmp/{data_name}_report.data {data_path}"
    hdcServer.execute_command_result(file_receive_cmd, shell=False)
    local_path = os.path.join(data_path, f"{data_name}_report.data")
    return local_path


def simpleperf_report(perf_data, report_file):
    my_path = os.path.dirname(os.path.realpath(__file__))
    py_path = os.path.join(my_path, "simpleperf", "report_html.py")
    result = subprocess.getoutput(f"python {py_path} -i {perf_data} -o {report_file} --no_browser")
    logger.info(f"result is: {result}")


def simpleperf_parse(mode, perf_data, api_name, html_name):  # 从perf文件获取指令数
    my_path = get_my_path()
    dir_path = os.path.join(my_path, "reports", mode, api_name)
    html_file_path = os.path.join(dir_path, f"{html_name}.html")
    if file_valid(perf_data):
        simpleperf_report(perf_data, html_file_path)
    cycles_keyword = '"cpu-cycles", "eventCount": '
    instructions_keyword = '"instructions", "eventCount": '
    if file_valid(html_file_path):
        with open(html_file_path, "r", encoding="utf-8") as f:
            content = f.read()
            cycles_count = re.search(r"{0}\d+".format(cycles_keyword), content).group().split(":")[1]
            instructions_count = re.search(r"{0}\d+".format(instructions_keyword), content).group().split(":")[1]
            logging.info(f"{api_name} cycles is: {cycles_count}, instructions is: {instructions_count}")

        return cycles_count, instructions_count


def three_sum_closest(nums, target):
    ret = float('inf')
    nums.sort()
    length = len(nums)
    for i in range(length - 2):
        left = i + 1
        right = length - 1
        while left < right:
            tmp = nums[i] + nums[left] + nums[right]
            ret = tmp if abs(tmp - target) < abs(ret - target) else ret
            if tmp == target:
                return target
            if tmp > target:
                right -= 1
            else:
                left += 1
    return ret


def api_test(hdcServer, demo_path, api_name, log_name, test_count=50, test_time=10, trace_name="", trace_time=10,
             # 单个用例测试
             data_name="",
             mode="drawing", log_time=15):
    delete_hilog(hdcServer)
    hdcServer.execute_command_result("rm -rf /data/local/tmp/*", shell=True)
    logger.info("xyj start thread")
    hilog_thread = MyThread(collecte_hilog, args=(hdcServer, api_name, log_name, mode, log_time))  # 日志落盘线程
    hilog_thread.start()
    demo_thread = MyThread(push_and_run_rs_demo,
                           args=(hdcServer, demo_path, api_name, test_count, test_time))  # 执行单个用例线程
    demo_thread.start()
    time.sleep(1)
    # yellow todo instruction todo
    # cycles_instructions_thread = MyThread(hiperf_collect,
    #                                       args=(hdcServer, mode, api_name, data_name, test_time)) #收集指令数线程
    # cycles_instructions_thread.start()
    hilog_thread.join()
    logger.info("xyj end hilog_thread")
    demo_thread.join()
    logger.info("xyj end demo_thread")
    # cycles_instructions_thread.join()
    # logger.info("xyj end cycles_instructions_thread")
    hilog_file_path = hilog_thread.get_result()
    # yellow todo instruction todo
    # perf_data = cycles_instructions_thread.get_result()
    # api_time, total_count = cal_api_time(api_name, hilog_file_path) #根据log打印计算执行时间、次数
    api_time = 10
    total_count = 1
    # yellow todo instruction todo
    # cycles_count, instructions_count = hiperf_parse(mode, perf_data, api_name, data_name) #根据perf获取指令数
    cycles_count = 100000
    instructions_count = 50000
    cycles = round(int(cycles_count) / total_count, 2)
    instructions = round(int(instructions_count) / total_count, 2)
    hdcServer.execute_command_result("rm -rf /data/local/tmp/*", shell=True)
    time.sleep(5)

    return api_time, total_count, cycles, instructions


def write_dict_to_json(result_dict, json_path):
    """
    :param result_dict:用于写入的数据字典
    :param json_path:json路径
    :return:无
    """
    json_str = json.dumps(result_dict)
    with open(json_path, "a", encoding='utf-8') as json_file:
        json_file.write(json_str + "\n")


class WriteExcel:
    def __init__(self, excel_path):
        self.ws = None
        self.wb = None
        self.excel_path = excel_path

    def create_excel(self):
        if file_valid(self.excel_path):
            os.remove(self.excel_path)
        # 创建工作表
        self.wb = openpyxl.Workbook()  # 创建工作薄
        self.wb.save(self.excel_path)
        self.wb.close()

    def write_data(self, data_list):
        with open(self.excel_path) as f:
            # 写入数据
            for data in data_list:
                values = []
                for va in data.values():
                    values.append(va)
                # 每次写入一行
                self.ws.append(values)
            self.adjust_format()
            self.wb.save(self.excel_path)
            self.wb.close()

    def adjust_format(self):
        table_header = []
        for column in self.ws.iter_cols(max_row=1, values_only=True):
            table_header.append(column)
        # 调整列宽
        for i in range(len(table_header)):
            k = get_column_letter(i + 1)
            column_list = []
            for j in range(1, self.ws.max_row + 1):
                column_list.append(len(str(self.ws.cell(row=j, column=i + 1).value)))
            col_width = max(column_list)
            self.ws.column_dimensions[k].width = max(col_width, len(table_header[i])) + 2

        # 定义表头颜色样式为蓝色
        header_fill = PatternFill("solid", fgColor="0099CCFF")
        # 定义对齐样式横向居中、纵向居中
        align = Alignment(horizontal="center", vertical="center")
        # 定义字体微软雅黑加粗
        font_size = 10
        font = Font("Microsoft YaHei", bold=False, size=font_size)
        for i in range(1, self.ws.max_row + 1):
            for j in range(1, self.ws.max_column + 1):
                if i == 1:
                    # 设置第一行单元格填充颜色
                    self.ws.cell(i, j).fill = header_fill
                    self.ws.cell(i, j).font = Font("Microsoft YaHei", bold=True, size=font_size)
                else:
                    # 设置字体
                    self.ws.cell(i, j).font = font
                # 设置单元格对其方式
                self.ws.cell(i, j).alignment = align
        self.wb.save(self.excel_path)
        self.wb.close()

    def open_create_sheet(self, sheet_name, index):
        self.create_excel()
        self.wb = openpyxl.load_workbook(self.excel_path)
        if file_valid(self.excel_path):
            with open(self.excel_path) as f:
                # 创建sheet表
                self.ws = self.wb.create_sheet(sheet_name, index)
                # 添加表头
                header = ["api_name", "api_time", "total_count", "instructions", "cycles"]
                self.ws.append(header)


# 读取测试列表
def read_api(sheet_name, model="drawing"):
    my_path = os.path.dirname(os.path.realpath(__file__))
    file_path = os.path.join(my_path, "oh_drawing_api.xlsx")
    data = pandas.read_excel(file_path, sheet_name=sheet_name)
    api_list = [api for api in data["api name"]]
    test_status = [status for status in data[f"{model} test status"]]
    api_test_list, api_finish_list = [], []
    for i in range(len(api_list)):
        if test_status[i] == "no":
            api_finish_list.append(api_list[i])
        if test_status[i] == "yes":
            api_test_list.append(api_list[i])
    logger.info(f"test api list is : {api_test_list}")
    return api_test_list


def continue_execute(hdcServer, mode, rs_demo_path, data_write_excel, test_api=None, test_count=50, test_time=10,
                     # 执行所有用例
                     trace_time=10, log_time=10, test_number=5):
    retry_api_list = []
    if test_api:
        for api_name in test_api:  # 遍历用例
            api_data_list, api_time_list, cycles_list, instructions_list, total_count_list = [], [], [], [], []
            failed_number = 0
            for i in range(test_number):
                trace_name = f"{mode}_{api_name}_{i + 1}"
                log_name = f"{mode}_{api_name}_{i + 1}"
                data_name = f"{mode}_{api_name}_{i + 1}"
                try:
                    api_time, total_count, cycles, instructions = api_test(hdcServer, rs_demo_path, api_name, log_name,
                                                                           # 单个用例执行
                                                                           test_count=test_count,
                                                                           test_time=test_time,
                                                                           trace_name=trace_name, trace_time=trace_time,
                                                                           data_name=data_name, mode=mode,
                                                                           log_time=log_time)
                    api_time_list.append(api_time)
                    cycles_list.append(cycles)
                    instructions_list.append(instructions)
                    total_count_list.append(total_count)
                    each_data = {"api_name": api_name, "api_time": api_time, "total_count": total_count,
                                 "instructions": instructions, "cycles": cycles}
                    api_data_list.append(each_data)
                    ave_data = {f"{api_name}_ave": api_name, "api_time": round(mean(api_time_list), 2),
                                "total_count": round(mean(total_count_list), 2),
                                "instructions": round(mean(instructions_list), 2),
                                "cycles": round(mean(cycles_list), 2)}
                except Exception as e:
                    logger.error(f"error is: {e}")
                    logger.error(f"{api_name} api test failed, please check!!!")
                    failed_number += 1
                    pass
                continue
            if failed_number >= 3:
                logger.error("=======================failed_number >= 3")
                retry_api_list.append(api_name)
            else:
                logger.error("=======================write_data")
                data_write_excel.write_data(api_data_list)  # 结果写入excel
                data_write_excel.adjust_format()
        logger.info(f"retry_api_list is: {retry_api_list}")

    return retry_api_list


def run_all_test(hdcServer, mode="drawing", test_api=None, test_count=50, test_time=10, trace_time=10, log_time=10,
                 test_number=5, retry_number=2, wait_time=60):
    """
    :param hdcServer: hdc 服务
    :param mode:测试模式, 分为skia和drawing
    :param test_api:测试api列表
    :param test_count:test_count的平方为每一帧绘制的次数
    :param test_time:测试时间
    :param trace_time:抓trace时间, 目前默认不抓取
    :param log_time:抓取hilog的时间
    :param test_number:每个测试次数, 测试5组, 取3组稳定数据的平均值
    :param retry_number: api测试失败后, 自动重新测试执行次数
    :param wait_time: 设备重启等待时间
    :return:
    """
    my_path = os.path.dirname(os.path.realpath(__file__))
    demo_path = os.path.join(my_path, "demo")
    result_path = os.path.join(my_path, "reports", mode)
    # yellow todo
    # push_simpleperf(hdcServer) #推进perf相关文件
    make_dir(result_path)  # 创建结果目录
    set_hilog(hdcServer)  # 设置hilog收集
    rs_demo_path = os.path.join(demo_path, mode, f"{mode}_rs_demo")
    time_now = time.strftime("%Y%m%d_%H%M%S", time.localtime())
    excel_path = os.path.join(result_path, f"{model}_api_perf_test_result_{time_now}.xlsx")
    data_write_excel = WriteExcel(excel_path)  # 创建excel对象
    data_write_excel.open_create_sheet(f"{model}_api_perf_test_result", 0)  # 创建表头
    retry_api_list = continue_execute(hdcServer, mode, rs_demo_path, data_write_excel, test_api=test_api,
                                      test_count=test_count, test_time=test_time, trace_time=trace_time,
                                      log_time=log_time, test_number=test_number)
    fail_api_list = []
    if retry_api_list and retry_number > 0:
        logger.info(f"Please wait for the device reboot.")
        logger.info(f"The 1st retry test api is: {retry_api_list}")
        for i in range(retry_number):
            reboot_and_unlock(hdcServer, wait_time=wait_time)  # 重启设备并解锁
            lock_core_frequency(hdcServer, clear_flag=False, wait_time=wait_time)  # 锁频锁核
            fail_api_list = continue_execute(hdcServer, mode, rs_demo_path, data_write_excel, test_api=test_api,
                                             test_count=test_count, test_time=test_time, trace_time=trace_time,
                                             log_time=log_time, test_number=test_number)  # 重新执行测试, 并验证是否存在问题
            logger.info(f"The {i + 1}st retry test api is: {fail_api_list}")
            if len(fail_api_list) == 0:
                logger.info("All api test finished!")
                break
        if len(fail_api_list) != 0:
            logger.error(f"Retry test finished, the failed api is: {fail_api_list}, please confirm!")


if __name__ == '__main__':
    hdc_tool = "hdc"
    # yellow todo
    # device = "sn"
    device = ""
    hdc_server = ExecuteHDC(hdc_tool, device)
    lock_core_frequency(hdc_server, clear_flag=False, wait_time=60)  # 锁频锁核
    # model值为skia或者drawing
    model = "drawing"
    test_api_list = read_api(sheet_name="oh_drawing_api", model=model)
    run_all_test(hdc_server, mode=model, test_api=test_api_list, test_count=10, test_time=10, trace_time=10,
                 log_time=30, test_number=5, retry_number=2, wait_time=60)
